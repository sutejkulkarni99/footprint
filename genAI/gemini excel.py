import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl

def extract_pins(input_file, output_file):
    try:
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        pin_row = None
        for row_index, row in enumerate(sheet.iter_rows()):
            for cell in row:
                if cell.value is not None and str(cell.value).strip().upper() == "PIN":
                    pin_row = row_index + 1
                    break
            if pin_row is not None:
                break

        if pin_row is None:
            raise ValueError("Could not find 'PIN' in the sheet.")

        pin_numbers = []
        for row_index in range(pin_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=1).value
            if cell_value is not None:
                try:
                    pin_number = float(cell_value)
                    pin_numbers.append(pin_number)
                except ValueError:
                    print(f"Skipping non-numeric value: {cell_value}") #Optional print statement for debugging
                    pass

        if not pin_numbers:
            raise ValueError("No numeric pin values found below 'PIN'.")

        df = pd.DataFrame({"Pin Number": pin_numbers})
        df.to_excel(output_file, index=False)
        return True
    except FileNotFoundError:
        return "Input file not found."
    except ValueError as e:
        return str(e)
    except Exception as e:
        return f"An unexpected error occurred: {e}"

def browse_input():
    filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    input_entry.delete(0, tk.END)
    input_entry.insert(0, filename)

def browse_output():
    filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    output_entry.delete(0, tk.END)
    output_entry.insert(0, filename)

def process_files():
    input_file = input_entry.get()
    output_file = output_entry.get()

    if not input_file or not output_file:
        messagebox.showerror("Error", "Please select both input and output files.")
        return

    result = extract_pins(input_file, output_file)

    if result is True:
        messagebox.showinfo("Success", f"Pin numbers extracted and saved to {output_file}")
    else:
        messagebox.showerror("Error", result)

# GUI Setup
window = tk.Tk()
window.title("Pin Extractor")

input_label = tk.Label(window, text="Input Excel File:")
input_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")

input_entry = tk.Entry(window, width=50)
input_entry.grid(row=0, column=1, padx=5, pady=5)

input_button = tk.Button(window, text="Browse", command=browse_input)
input_button.grid(row=0, column=2, padx=5, pady=5)

output_label = tk.Label(window, text="Output Excel File:")
output_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")

output_entry = tk.Entry(window, width=50)
output_entry.grid(row=1, column=1, padx=5, pady=5)

output_button = tk.Button(window, text="Browse", command=browse_output)
output_button.grid(row=1, column=2, padx=5, pady=5)

process_button = tk.Button(window, text="Extract Pins", command=process_files)
process_button.grid(row=2, column=1, pady=10)

window.mainloop()