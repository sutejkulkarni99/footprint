import sys
import openpyxl
from openpyxl.utils import range_boundaries
import xml.etree.ElementTree as ET
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                               QPushButton, QLabel, QLineEdit, QFileDialog, 
                               QTextEdit, QMessageBox)
from xml.dom import minidom

class ExcelToXmlConverter(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel to XML Converter")
        self.setGeometry(100, 100, 800, 600)
        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        layout = QVBoxLayout()

        # File Selection
        self.file_path = QLineEdit()
        browse_btn = QPushButton("Browse Excel File")
        browse_btn.clicked.connect(self.browse_file)

        # Convert Button
        convert_btn = QPushButton("Convert to XML")
        convert_btn.clicked.connect(self.convert)

        # Status Display
        self.status = QTextEdit()
        self.status.setReadOnly(True)

        layout.addWidget(QLabel("Excel File:"))
        layout.addWidget(self.file_path)
        layout.addWidget(browse_btn)
        layout.addWidget(convert_btn)
        layout.addWidget(self.status)

        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

    def browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)")
        if path:
            self.file_path.setText(path)

    def get_merged_regions(self, sheet):
        """Handle merged cells in read-only mode"""
        merged_regions = []
        for merge_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(merge_range)
            if min_row == 1 and max_row == 1:  # Only first row merges
                merged_regions.append((min_col, max_col))
        return merged_regions

    def process_sheet(self, ws):
        """Process a worksheet and return cable data"""
        try:
            merged_regions = self.get_merged_regions(ws)
            connectors = []

            # Identify connectors and their pin columns
            for min_col, max_col in merged_regions:
                connector_name = ws.cell(row=1, column=min_col).value
                if not connector_name:
                    continue

                # Find pin column under the connector
                pin_col = None
                for col in range(min_col, max_col + 1):
                    header = ws.cell(row=2, column=col).value
                    if header and "pin" in str(header).lower():
                        pin_col = col
                        break

                if pin_col:
                    connectors.append({
                        'name': connector_name.strip(),
                        'pin_col': pin_col - 1,  # Convert to 0-based index
                        'pins': set()
                    })

            # Process connections
            connections = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                current_pins = {}
                for conn in connectors:
                    pin_value = row[conn['pin_col']]
                    if isinstance(pin_value, (int, float)) and pin_value > 0:
                        pin = int(pin_value)
                        current_pins[conn['name']] = pin
                        conn['pins'].add(pin)

                # Create all possible connections in this row
                conn_names = list(current_pins.keys())
                for i in range(len(conn_names)):
                    for j in range(i + 1, len(conn_names)):
                        connections.append({
                            'from': f"{conn_names[i]}:{current_pins[conn_names[i]]}",
                            'to': f"{conn_names[j]}:{current_pins[conn_names[j]]}",
                            'type': "Wire"
                        })

            # Calculate max pins
            for conn in connectors:
                conn['max_pin'] = max(conn['pins']) if conn['pins'] else 0

            return {
                'cable_name': f"Cable_{ws.title}",
                'connectors': connectors,
                'connections': connections
            }

        except Exception as e:
            raise ValueError(f"Error processing sheet: {str(e)}")

    def convert(self):
        try:
            if not self.file_path.text():
                raise ValueError("Please select an Excel file")

            wb = openpyxl.load_workbook(self.file_path.text(), read_only=True)
            root = ET.Element("CableList")

            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    cable_data = self.process_sheet(ws)

                    # Create cable element
                    cable = ET.SubElement(root, "Cable", Name=cable_data['cable_name'])

                    # Add connectors
                    connectors_elem = ET.SubElement(cable, "Connectors")
                    for conn in cable_data['connectors']:
                        connector = ET.SubElement(
                            connectors_elem, "Connector",
                            Name=conn['name'],
                            ConName=conn['name'],
                            ConID=conn['name'].replace(" ", "")
                        )
                        ET.SubElement(connector, "Pins").text = str(conn['max_pin'])

                    # Add connections
                    fromto = ET.SubElement(cable, "FromTo")
                    for connection in cable_data['connections']:
                        ET.SubElement(fromto, "Cx",
                                    From=connection['from'],
                                    To=connection['to'],
                                    Type=connection['type'])

                    self.status.append(f"Processed sheet: {sheet_name}")

                except Exception as e:
                    self.status.append(f"Skipped sheet {sheet_name}: {str(e)}")

            # Save XML
            xml_str = ET.tostring(root, encoding='unicode')
            dom = minidom.parseString(xml_str)
            pretty_xml = dom.toprettyxml(indent="    ")

            save_path, _ = QFileDialog.getSaveFileName(
                self, "Save XML File", "", "XML Files (*.xml)")

            if save_path:
                if not save_path.endswith('.xml'):
                    save_path += '.xml'
                with open(save_path, 'w') as f:
                    f.write(pretty_xml)
                self.status.append(f"XML saved to: {save_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            self.status.append(f"Fatal error: {str(e)}")
        finally:
            if 'wb' in locals():
                wb.close()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    converter = ExcelToXmlConverter()
    converter.show()
    sys.exit(app.exec())